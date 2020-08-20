#!/usr/bin/env python3

import argparse
import json

from io import open
from termcolor import colored, cprint
from slugify import slugify
from bson.objectid import ObjectId
from openpyxl import load_workbook
from datetime import *

def check_arguments():
    """Parse the call to the CLI

    Returns:
        namespace -- [Namespaces for all the parameters passed to the CLI]
    """
    parse = argparse.ArgumentParser(prog='import-to-json',
                                    usage='%(prog)s [-h|--help] -f|--file (file.xls) -j|--json input.json -s|--swimlane name [-o|--output output.json]',
                                    description='Add cards read from an external file to a JSON file exported from Wekan',
                                    epilog='',
                                    allow_abbrev=False)
    parse.add_argument('-f','--file',
                        type=str,
                        help='File from where new cards are loaded',
                        required=True)    
                        
    parse.add_argument('-j','--json',
                        type=str,
                        help='JSON file exported from Wekan',
                        required=True)    
                        
    parse.add_argument('-s','--swimlane',
                        type=str,
                        help='Swimlane containing all imported cards, which may or not may exist',
                        required=True)

    parse.add_argument('-o','--output',
                        help='(Optional) Final JSON with all cards joined from both files',
                        type=str)

    args = parse.parse_args()
    return args


def load_from_json(file):
    """Read and load the json file context

    Arguments:
        file {str} -- Absolute or relative path to json file exported from Wekan

    Returns:
        dict -- Data structure corresponding to the content of the json file
    """
    with open(file, 'r') as filePointer:
        context = filePointer.read()
        jsonText = json.loads(context)
        filePointer.close()

    return jsonText


def json_summary(jsonDict):
    """Prints a summary content of the json file loaded 

    Arguments:
        jsonDict {dict} -- Dictionary corresponding to the content of the json file
    """
    cprint('\nThe json file have been imported correctly!', 'green', 'on_white',attrs=['bold'])
    print ("Brief summary:")
    print ("\tBoard imported: ", jsonDict['title'])
    print ("\tTotal quantity of lists imported: ", len(jsonDict['lists']))
    print ("\tTotal quantity of users imported: ", len(jsonDict['users']))
    print ("\tTotal quantity of swimlanes imported: ", len(jsonDict['swimlanes']))
    print ("\tTotal quantity of cards imported: ", len(jsonDict['cards']))


def write_to_json(file, jsonDict):
    """Writing a dictionary to a json file

    Arguments:
        file {str} -- Absolute or relative path to the json file
        jsonDict {dict} -- Dictionary with the information to save in the json file
    """
    with open(file,'w') as filePointer:
        filePointer.write(json.dumps(jsonDict, ensure_ascii=False))
        filePointer.close()


def there_is_user(userList, user_slug):
    """Find if a user belong to the board

    Arguments:
        userList {list} -- list of users belonging to the imported json file
        user_slug {str} -- User slugify 

    Returns:
        str -- User ID or None
    """
    for user in userList:
        if user_slug in slugify(user['profile']['fullname'],to_lower=True):
            return user['_id']
    return None


def there_is_swimlane(swimlaneList,swimlane_slug):
    """Find if a swimlane belong to the board

    Arguments:
        swimlaneList {list} -- list of swimlanes belonging to the imported json file
        swimlane_slug {str} -- Swimlane slugify

    Returns:
        str -- Swimlane ID or None
    """
    for swimlane in swimlaneList:
        if swimlane_slug in slugify(swimlane['title'],to_lower=True):
            return swimlane['_id']
    return None

def create_new_swimlane(title):
    """Create a dictionary representing a new swimlane

    Arguments:
        title {str} -- Swimlane name or title

    Returns:
        dict -- A new swimlane dictionary
    """
    return dict(_id=str(ObjectId()), title=title, archived=False, type='swimlane' )    


def create_new_card(title, description, assignees, startAt, dueAt, listId, labelIds , customFields, swimlaneId)  :
    return dict(
                _id=str(ObjectId()), 
                title=title, 
                description=description, 
                assignees= list(assignees), 
                startAt= startAt, 
                dueAt= dueAt, 
                listId= listId, 
                swimlaneId= swimlaneId,
                labelIds= list(labelIds),
                customFields= list(customFields),
                archived= False
                )


def read_from_excel(pathfile):
    """Read all rows from the first sheet of the document

    Arguments:
        pathfile {str} -- Absolute or relative path to the excel file

    Yields:
        iterator -- Iterator representing all file rows
    """
    workbook = load_workbook(filename=pathfile)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=1, max_col=6, values_only=True):
        yield row
        

def main():
    """Main function which orchestrate the script
    """
    error = False
    try:
        args = check_arguments() # Checking that all needed parameters are right
        jsonText = load_from_json(args.json) # Loading the json file
        json_summary(jsonText) # Printing a summary of the json file loaded previously

        if args.output == None: 
            print(json.dumps(jsonText, ensure_ascii=False, indent=4)) # Print to stdout the final JSON   
        else:
            write_to_json(args.output, jsonText) # Write to a Json file the result

        xlsIterator = read_from_excel(args.file)
        next(xlsIterator); next(xlsIterator)
        for row in xlsIterator:
            print (row)

    except FileNotFoundError:
        print ("Error: Json file {} not found".format(args.json))
        error = True
    except PermissionError:
        cprint ("Denied permission: Do not have permission to create a json file: {}".format(args.output), 'red', attrs=['bold'])
        error = True
    finally:
        if error:
            cprint("Script execution ends with error!",'red','on_white',attrs=['bold'])


if __name__ == "__main__":
    main()